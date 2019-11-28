import openpyxl


def getRowCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(SheetName)
    return (sheet.max_row)

def getColumnCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(SheetName)
    return (sheet.max_column)

def ReadData(file,SheetName,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(SheetName)
    return sheet.cell(row=rownum,column=colnum).value

def WriteData(file,Sheetname,rownum,colnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(Sheetname)
    sheet.cell(row=rownum,column=colnum).value= data
    workbook.save(file)


