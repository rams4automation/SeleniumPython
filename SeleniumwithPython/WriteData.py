
import openpyxl

path="C:\TestData\Python1.xlsx"

workbook=openpyxl.load_workbook(path)

sheet=workbook.active

for r in range(1,6):
    for c in range(1,3):
        if r==1:
         sheet.cell(row=1,column=1).value="UserName"
         sheet.cell(row=1,column=2).value="Password"
        elif c==2:
            sheet.cell(row=r, column=2).value = "Password1"
        else:
            sheet.cell(row=r,column=c).value="Welcome"

workbook.save(path)

workbook.close()