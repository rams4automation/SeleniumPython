
import openpyxl

def getrowcount(filepath,sheetName):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet=workbook.active
    rows=sheet.max_row
    return rows

def getcolumnscount(filepath,sheetName):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet=workbook.active
    cols=sheet.max_column
    return cols

def getTestData(filepath,sheetName,rownum,colnum):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet=workbook.active
    strdata=sheet.cell(row=rownum,column=colnum).value
    return strdata

def writedata(filepath,sheetName,rownum,colnum,data):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet=workbook.active
    sheet.cell(row=rownum,column=colnum).value=data
    workbook.save(filepath)



